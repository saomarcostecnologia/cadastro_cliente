from typing import Optional, Tuple, Union
import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook, workbook
from tkcalendar import DateEntry
from dateutil.relativedelta import relativedelta
from datetime import datetime
from tkinter import Toplevel
import os
from tkinter import filedialog
import platform

# Setando a aparencia do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class EditDialog(Toplevel):
    def __init__(self, parent, client_data):
        super().__init__(parent)
        self.title("Editar Cliente")
        self.geometry("400x300")
        self.transient(parent)

        self.client_data = client_data

        self.lb_name = Label(self, text="Nome Completo:")
        self.name_entry = Entry(self, width=40)
        self.lb_name.grid(row=0, column=0, padx=10, pady=10)
        self.name_entry.grid(row=0, column=1, padx=10, pady=10)
        self.name_entry.insert(0, client_data[0])  # Set the initial value

        self.lb_phone = Label(self, text="Telefone:")
        self.phone_entry = Entry(self, width=20)
        self.lb_phone.grid(row=1, column=0, padx=10, pady=10)
        self.phone_entry.grid(row=1, column=1, padx=10, pady=10)
        self.phone_entry.insert(0, client_data[1])  # Set the initial value

        # Add similar widgets for other client attributes (age, gender, plan, etc.)

        self.save_button = Button(self, text="Salvar", command=self.save_changes)
        self.save_button.grid(row=10, columnspan=2, pady=20)

    
    def save_changes(self):
        new_name = self.name_entry.get()
        new_phone = self.phone_entry.get()

        updated_client_data = (new_name, new_phone, *self.client_data[2:])

        ficheiro = openpyxl.load_workbook("Clientes.xlsx")
        folha = ficheiro.active

        rows = list(folha.iter_rows(min_row=2, values_only=True))  # Lê as linhas em uma lista

        for idx, row in enumerate(rows):
            if row[1] == self.client_data[1]:
                rows[idx] = updated_client_data  # Substitui a tupla inteira pela nova lista

        # Limpa o conteúdo da folha
        for row in folha.iter_rows(min_row=2, max_row=folha.max_row, min_col=1, max_col=len(self.client_data)):
            for cell in row:
                cell.value = None

        # Preenche a folha com as novas linhas atualizadas
        for idx, updated_row in enumerate(rows, start=2):
            for col_idx, value in enumerate(updated_row, start=1):
                folha.cell(row=idx, column=col_idx, value=value)

        ficheiro.save("Clientes.xlsx")
        messagebox.showinfo("Cliente Atualizado", "As informações do cliente foram atualizadas com sucesso.")
        self.destroy()

class SearchDialog(Toplevel):
    def __init__(self, parent, search_callback):
        super().__init__(parent)
        self.title("Localizar Cliente")
        self.geometry("400x150")
        self.transient(parent)

        self.label = Label(self, text="Digite o número de telefone do cliente:")
        self.label.pack(pady=10)

        self.phone_entry = Entry(self)
        self.phone_entry.pack()

        self.search_button = Button(self, text="Buscar", command=lambda: search_callback(self.phone_entry.get()))
        self.search_button.pack(pady=10)

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.apperence()
        self.todo_sistema()

    def layout_config(self):
        self.title("Sistema de Gestão de Clientes - Nutricionista Karina Szeibl")
        self.geometry("700x500")
        self.update_idletasks()  # Calcular tamanho da janela antes de centralizar
        width = self.winfo_width()
        height = self.winfo_height()
        x_offset = (self.winfo_screenwidth() - width) // 2
        y_offset = (self.winfo_screenheight() - height) // 2
        self.geometry(f"{width}x{height}+{x_offset}+{y_offset}")

    def apperence(self):
        self.lb_apm = ctk.CTkLabel(
            self, text="Tema", bg_color="transparent", text_color=["#000", "#fff"]
        ).place(x=50, y=440)
        self.opt_apm = ctk.CTkOptionMenu(
            self, values=["Light", "Dark", "System"], command=self.change_apm
        ).place(x=50, y=465)
    
    def open_search_dialog(self):
        search_dialog = SearchDialog(self, self.find_client)
        self.wait_window(search_dialog)

    def find_client(self, phone):
        ficheiro = openpyxl.load_workbook("Clientes.xlsx")
        folha = ficheiro.active

        for row in folha.iter_rows(min_row=2, values_only=True):
            if row[1] == phone:  # Assuming phone number is in the second column
                edit_dialog = EditDialog(self, row)  # Pass the client data to the edit dialog
                self.wait_window(edit_dialog)
                return

        messagebox.showinfo("Cliente não encontrado", "Nenhum cliente encontrado com este número de telefone.")


    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)
    
    def todo_sistema(self):
        frame = ctk.CTkFrame(
            self,
            width=700,
            height=50,
            corner_radius=0,
            bg_color="teal",
            fg_color="teal",
        ).place(x=0, y=10)
        title = ctk.CTkLabel(
            frame,
            text="Sistema de Gestão de Clientes",
            font=("Century Gothic", 24),
            text_color="#fff",
            bg_color="teal",
        ).place(x=170, y=20)
        span = ctk.CTkLabel(
            self,
            text="Por Favor, preencha todos os campos do formulário!",
            font=("Century Gothic", 16),
            text_color=["#000", "#fff"],
        ).place(x=50, y=70)

        ficheiro = pathlib.Path("Clientes.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro = Workbook()
            folha = ficheiro.active
            folha["A1"] = "Nome Completo"
            folha["B1"] = "Telefone"
            folha["C1"] = "Idade"
            folha["D1"] = "CPF"
            folha["E1"] = "Genero"
            folha["F1"] = "Endereço"
            folha["G1"] = "Email"
            folha["H1"] = "Plano"
            folha["I1"] = "Data Inicio"
            folha["J1"] = "Data Fim"
            folha["K1"] = "Observações"

            ficheiro.save("Clientes.xlsx")

        # Funções
        def submit():
            # pegando os dados dos entrys
            name = name_value.get()
            phone = phone_value.get()
            age = age_value.get()
            cpf = cpf_value.get()
            gender = gender_combobox.get()
            plan = plan_combobox.get()
            dt_inicio = dt_inicio_value.get()
            adress = adress_value.get()
            email = email_value.get()
            obs = obs_value.get()           

            if (
                name == "" or phone == "" or email == ""
            ):  ## Criar validação para vazio e existente.
                messagebox.showerror(
                    "Sistema",
                    "Operação não concluida!\nPor favor Preencha todos os campos",
                )
            else:
                ficheiro = openpyxl.load_workbook("Clientes.xlsx")
                folha = ficheiro.active
                folha.cell(column=1, row=folha.max_row + 1, value=name)
                folha.cell(column=2, row=folha.max_row, value=phone)
                folha.cell(column=3, row=folha.max_row, value=age)
                folha.cell(column=4, row=folha.max_row, value=cpf)
                folha.cell(column=5, row=folha.max_row, value=gender)
                folha.cell(column=6, row=folha.max_row, value=adress)
                folha.cell(column=7, row=folha.max_row, value=email)
                folha.cell(column=8, row=folha.max_row, value=plan)
                folha.cell(column=9, row=folha.max_row, value=dt_inicio)
                folha.cell(column=11, row=folha.max_row, value=obs)

                dt_inicio = dt_inicio_value.get()
                try:
                    dt_inicio_date = datetime.strptime(dt_inicio, "%d/%m/%Y")
                except ValueError:
                    messagebox.showerror(
                        "Erro de Data",
                        "Formato de data inválido para Data Inicio. Use o formato DD/MM/AAAA.",
                    )
                    return

                dt_inicio_date = datetime.strptime(dt_inicio, "%d/%m/%Y")  # Converter para objeto de data e hora
                plan_durations = {
                    "Mensal": 1,
                    "Trimestral": 3,
                    "Semestral": 6,
                    "Anual": 12,
                    "Básico": 1  # Você precisa determinar a duração correta para o plano "Básico"
                }
                selected_plan_duration = plan_durations[plan]
                
                # Calcular data de término usando a relativedelta
                dt_fim_date = dt_inicio_date + relativedelta(months=selected_plan_duration)
                dt_fim = dt_fim_date.strftime("%d/%m/%Y")  # Converter de volta para string

                folha.cell(column=10, row=folha.max_row, value=dt_fim)
                

                ficheiro.save(r"Clientes.xlsx")
                messagebox.showinfo("Sistema", "Dados Salvos com Sucesso!")
                clear()

        def clear():
            name_value.set("")
            phone_value.set("")
            age_value.set("")
            adress_value.set("")
            email_value.set("")
            cpf_value.set("")
            obs_value.set("")
            dt_inicio_value.set("")

        # test variables
        name_value = StringVar()
        phone_value = StringVar()
        age_value = StringVar()
        adress_value = StringVar()
        email_value = StringVar()
        cpf_value = StringVar()
        obs_value = StringVar()
        dt_inicio_value = StringVar()

        # Entrys
        name_entry = ctk.CTkEntry(
            self,
            textvariable=name_value,
            width=390,
            font=("Century Gothic", 16),
            fg_color="transparent",
        )
        contact_entry = ctk.CTkEntry(
            self,
            textvariable=phone_value,
            width=150,
            font=("Century Gothic", 16),
            fg_color="transparent",
        )
        age_entry = ctk.CTkEntry(
            self,
            textvariable=age_value,
            width=140,
            font=("Century Gothic", 16),
            fg_color="transparent",
        )
        cpf_entry = ctk.CTkEntry(
            self,
            textvariable=cpf_value,
            width=150,
            font=("Century Gothic", 16),
            fg_color="transparent",
        )
        adress_entry = ctk.CTkEntry(
            self,
            textvariable=adress_value,
            width=200,
            font=("Century Gothic", 16),
            fg_color="transparent",
        )
        email_entry = ctk.CTkEntry(
            self,
            textvariable=email_value,
            width=200,
            font=("Century Gothic", 16),
            fg_color="transparent",
        )
        # Combo box
        gender_combobox = ctk.CTkComboBox(
            self,
            values=["Masculino", "Feminino", "Outro"],
            font=("Century Gothic", 14),
            state="readonly", 
        )
        gender_combobox.set("Feminino")
        # Combo box plano
        plan_combobox = ctk.CTkComboBox(
            self,
            values=["Mensal", "Trimestral", "Semestral", "Anual", "Básico"],
            font=("Century Gothic", 14),
            state="readonly", 
            width= 150
        )
        plan_combobox.set("Mensal")

        # Entrada de observações
        obs_entry = ctk.CTkEntry(
            self,
            textvariable=obs_value,
            width=398,
            height=75,
            font=("arial", 12),
            border_color="#aaa",
            border_width=2,
            fg_color="transparent",
        )
        dt_inicio_entry = ctk.CTkEntry(
            self,
            textvariable=dt_inicio_value,
            width=150,
            font=("Century Gothic", 16),
            fg_color="transparent",
        )

        # Botoes submit e limpar
        button_submit = ctk.CTkButton(
            self,
            text="Salvar dados".upper(),
            command=submit,
            fg_color="#151",
            hover_color="#131",
        ).place(x=375, y=465)
        button_clear = ctk.CTkButton(
            self,
            text="Limpar Campos".upper(),
            command=clear,
            fg_color="#555",
            hover_color="#333",
        ).place(x=525, y=465)

        button_find = ctk.CTkButton(
            self,
            text="Localizar Cadastro".upper(),
            command=self.open_search_dialog,
            fg_color="#222",
            hover_color="#111",
        ).place(x=210, y=465)

        # Labels
        lb_name = ctk.CTkLabel(
            self,
            text="Nome Completo:",
            font=("Century Gothic", 16),
            text_color=["#000", "#fff"],
        )
        lb_contact = ctk.CTkLabel(
            self,
            text="Contato:",
            font=("Century Gothic", 16),
            text_color=["#000", "#fff"],
        )
        lb_age = ctk.CTkLabel(
            self,
            text="Data de nascimento:",
            font=("Century Gothic", 16),
            text_color=["#000", "#fff"],
        )
        lb_cpf = ctk.CTkLabel(
            self,
            text="CPF:",
            font=("Century Gothic", 16),
            text_color=["#000", "#fff"],
        )
        lb_gender = ctk.CTkLabel(
            self,
            text="Gênero:",
            font=("Century Gothic", 16),
            text_color=["#000", "#fff"],
        )
        lb_adress = ctk.CTkLabel(
            self,
            text="Endereço:",
            font=("Century Gothic", 16),
            text_color=["#000", "#fff"],
        )
        lb_plan = ctk.CTkLabel(
            self,
            text="Plano:",
            font=("Century Gothic", 16),
            text_color=["#000", "#fff"],
        )
        lb_email = ctk.CTkLabel(
            self,
            text="Email:",
            font=("Century Gothic", 16),
            text_color=["#000", "#fff"],
        )
        lb_obs = ctk.CTkLabel(
            self,
            text="Observações:",
            font=("Century Gothic", 16),
            text_color=["#000", "#fff"],
        )
        lb_dt_inicio = ctk.CTkLabel(
            self,
            text="Data Inicio:",
            font=("Century Gothic", 16),
            text_color=["#000", "#fff"],
        )

        # Posicionando na tela
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contact.place(x=500, y=120)
        contact_entry.place(x=500, y=150)

        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)

        lb_cpf.place(x=500, y=190)
        cpf_entry.place(x=500,y=220)

        lb_gender.place(x=300, y=260)
        gender_combobox.place(x=300, y=290)

        lb_plan.place(x=500, y=260)
        plan_combobox.place(x=500, y=290)

        lb_adress.place(x=50, y=260)
        adress_entry.place(x=50, y=290)

        lb_email.place(x=50, y=190)
        email_entry.place(x=50, y=220)

        lb_obs.place(x=50, y=330)
        obs_entry.place(x=50, y=360)

        lb_dt_inicio.place(x=500, y=330)
        dt_inicio_entry.place(x=500, y=360)


if __name__ == "__main__":
    app = App()
    app.mainloop()